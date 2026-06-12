#!/usr/bin/env python3

import sys
import os
import shutil
import argparse
from openpyxl import load_workbook
from pccf_utils import get_hoja_label


def main():
    parser = argparse.ArgumentParser(
        description="Prepara el Excel del docente: renombra hojas a códigos cortos "
                    "y lo mueve a excels_{familia}/libro_{ciclo}.xlsx"
    )
    parser.add_argument("-c", "--ciclo", required=True,
                        help="Ciclo formativo (DAM, DAW, ASIR, SMX, ...)")
    parser.add_argument("-f", "--familia", required=True,
                        help="Familia profesional (INF, SCO, ...)")
    parser.add_argument("-i", "--input",
                        help="Ruta al Excel modificado por el docente "
                             "(por defecto: PDFS/libro_autogenerado_{CICLO}.xlsx)")
    parser.add_argument("--no-backup", action="store_true",
                        help="No crear copia de respaldo del destino si existe")

    args = parser.parse_args()

    ciclo = args.ciclo.upper()
    familia = args.familia.upper()

    if args.input:
        ruta_origen = args.input
    else:
        ruta_origen = f"PDFS/libro_autogenerado_{ciclo}.xlsx"

    if not os.path.exists(ruta_origen):
        print(f"ERROR: No se encuentra el archivo: {ruta_origen}")
        print(f"Ejecuta primero 'make proyecto-{ciclo}' para generarlo.")
        sys.exit(1)

    dir_destino = f"excels_{familia}"
    ruta_destino = f"{dir_destino}/libro_{ciclo}.xlsx"

    if not os.path.exists(dir_destino):
        print(f"Creando directorio: {dir_destino}/")
        os.makedirs(dir_destino)

    if os.path.exists(ruta_destino) and not args.no_backup:
        backup = f"{ruta_destino}.bak"
        n = 1
        while os.path.exists(backup):
            backup = f"{ruta_destino}.bak.{n}"
            n += 1
        print(f"  Backup: {ruta_destino} → {backup}")
        shutil.copy2(ruta_destino, backup)

    print(f"  Leyendo: {ruta_origen}")
    wb = load_workbook(ruta_origen)
    nombres_originales = wb.sheetnames

    cambios = []
    for ws in wb.worksheets:
        nombre_original = ws.title
        nombre_nuevo = get_hoja_label(nombre_original)
        if nombre_nuevo != nombre_original:
            ws.title = nombre_nuevo
            cambios.append((nombre_original, nombre_nuevo))
            print(f"  Hoja: '{nombre_original}' → '{nombre_nuevo}'")
        else:
            print(f"  Hoja: '{nombre_original}' (sin cambios)")

    print(f"  Guardando: {ruta_destino}")
    wb.save(ruta_destino)
    print(f"\n✅ Excel preparado: {ruta_destino}")
    print(f"   Total hojas: {len(nombres_originales)}, renombradas: {len(cambios)}")


if __name__ == "__main__":
    main()
