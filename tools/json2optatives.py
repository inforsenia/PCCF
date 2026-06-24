#!/usr/bin/env python3

import sys
import json
import os
import argparse
from box import Box
import jinja2
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, NamedStyle
import warnings
warnings.filterwarnings('ignore')

from pccf_utils import OPTATIVES_PATH, OPTATIVES_PLANTILLES, PROJECT_DIR

numberStyle = NamedStyle(name='numberStyle', number_format='0.00')

parser = argparse.ArgumentParser(description="Genera Excel i PDs per als mòduls optatius compartits")
parser.add_argument("--outdir", help="Directori d'eixida (defecte: optatives/)")
args = parser.parse_args()

outdir = args.outdir or os.path.join(PROJECT_DIR, "optatives")
libro_path = os.path.join(outdir, "libro_optatives.xlsx")
plantilles_dir = os.path.join(outdir, "plantilles")

os.makedirs(plantilles_dir, exist_ok=True)

with open(OPTATIVES_PATH, encoding='utf-8') as f:
    optatives = json.load(f)

data_box = Box(optatives)

# -----------------------------------------------------------------------
# 1. Generate shared optatives Excel
# -----------------------------------------------------------------------
wb = openpyxl.Workbook()

p_codigo = 'C1'
p_nombre = 'C2'
p_horas = 'C3'
p_ra_col_l = 'B'
p_ra_titulo_col = 2
p_ra_titulo_row = 8
p_ce_col_l = 'E'
p_req_fe_col_l = 'H'
p_contenidos_col_l = 'J'
p_TOTAL_HORAS_titulo = "F3"
p_TOTAL_HORAS = "F5"
p_TOTAL_HORAS_DUAL_titulo = "I3"
p_TOTAL_HORAS_DUAL = "I5"

for codigo, modulo in data_box.items():
    print(f" [ Optatives Excel ] : {modulo.nombre}")
    ws = wb.create_sheet(title=modulo.nombre)
    wb.active = wb.sheetnames.index(modulo.nombre)

    # Reset row counters per module (same pattern as json2excel.py)
    p_ra_titulo_col = 2
    p_ra_titulo_row = 8
    p_ce_col = 5
    p_ce_row = 8
    p_comp_col = 4
    p_comp_row = 8
    p_h_col = 6
    p_h_row = 8
    p_ce_per_col = 7
    p_ce_per_row = 8
    p_req_fe_col = 8
    p_req_fe_row = 8
    p_horas_dual_col = 9
    p_horas_dual_row = 8
    p_contenidos_col = 10
    p_contenidos_row = 8

    ws['B1'].value = "Código"
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B1'].fill = PatternFill('lightHorizontal')

    ws['B2'].value = "Nombre"
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B2'].fill = PatternFill('lightHorizontal')

    ws['B3'].value = "Horas"
    ws['B3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B3'].fill = PatternFill('lightHorizontal')

    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=5)
    ws['C3'].value = modulo.horas
    ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C3'].fill = PatternFill('darkTrellis')
    ws['C3'].font = Font(size=13)

    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=5)
    ws[p_codigo].value = codigo
    ws[p_codigo].alignment = Alignment(horizontal='center', vertical='center')
    ws[p_codigo].fill = PatternFill('darkTrellis')
    ws[p_codigo].font = Font(size=13)

    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=5)
    ws[p_nombre].value = modulo.nombre
    ws[p_nombre].alignment = Alignment(horizontal='center', vertical='center')
    ws[p_nombre].fill = PatternFill('darkTrellis')
    ws[p_nombre].font = Font(size=14)

    ws.merge_cells(start_row=3, start_column=6, end_row=4, end_column=6)
    ws[p_TOTAL_HORAS_titulo].value = "TOTAL HORAS"
    ws[p_TOTAL_HORAS_titulo].fill = PatternFill('darkTrellis')
    ws[p_TOTAL_HORAS_titulo].font = Font(size=13)
    ws[p_TOTAL_HORAS_titulo].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.column_dimensions['I'].width = 15
    ws[p_TOTAL_HORAS].font = Font(size=16)
    ws[p_TOTAL_HORAS].value = "=SUM(F8:F200)/2"

    ws.merge_cells(start_row=3, start_column=9, end_row=4, end_column=9)
    ws[p_TOTAL_HORAS_DUAL_titulo].value = "TOTAL H.DUAL"
    ws[p_TOTAL_HORAS_DUAL_titulo].fill = PatternFill('darkTrellis')
    ws[p_TOTAL_HORAS_DUAL_titulo].font = Font(size=13)
    ws[p_TOTAL_HORAS_DUAL_titulo].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws[p_TOTAL_HORAS_DUAL].font = Font(size=14)
    ws.column_dimensions['F'].width = 15
    ws[p_TOTAL_HORAS_DUAL].value = "=SUM(I8:I200)/2"

    ws.merge_cells(start_row=p_ra_titulo_row, start_column=p_ra_titulo_col,
                   end_row=p_ra_titulo_row + 1, end_column=p_ra_titulo_col)
    ws.cell(column=p_ra_titulo_col, row=p_ra_titulo_row).value = "RESULTADO DE APRENDIZAJE"
    ws.cell(column=p_ra_titulo_col, row=p_ra_titulo_row).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(column=p_ra_titulo_col, row=p_ra_titulo_row).fill = PatternFill('gray125')
    ws.column_dimensions[p_ra_col_l].width = 40

    p_percent_ra_col = p_ra_titulo_col + 1
    p_percent_ra_row = p_ra_titulo_row
    ws.merge_cells(start_row=p_percent_ra_row, start_column=p_percent_ra_col,
                   end_row=p_percent_ra_row + 1, end_column=p_percent_ra_col)
    ws.cell(column=p_percent_ra_col, row=p_percent_ra_row).value = "% RA"
    ws.cell(column=p_percent_ra_col, row=p_percent_ra_row).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(column=p_percent_ra_col, row=p_percent_ra_row).fill = PatternFill('gray125')

    p_comp_col = p_percent_ra_col + 1
    p_comp_row = p_percent_ra_row
    ws.merge_cells(start_row=p_comp_row, start_column=p_comp_col,
                   end_row=p_comp_row + 1, end_column=p_comp_col)
    ws.cell(column=p_comp_col, row=p_comp_row).value = "COMP"
    ws.cell(column=p_comp_col, row=p_comp_row).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(column=p_comp_col, row=p_comp_row).fill = PatternFill('gray125')

    p_ce_col = p_comp_col + 1
    p_ce_row = p_comp_row
    ws.merge_cells(start_row=p_ce_row, start_column=p_ce_col,
                   end_row=p_comp_row + 1, end_column=p_ce_col)
    ws.cell(column=p_ce_col, row=p_ce_row).value = "CRITERIOS DE EVALUACIÓN"
    ws.cell(column=p_ce_col, row=p_ce_row).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(column=p_ce_col, row=p_ce_row).fill = PatternFill('gray125')
    ws.column_dimensions[p_ce_col_l].width = 90

    p_h_col = p_ce_col + 1
    p_h_row = p_ce_row
    ws.merge_cells(start_row=p_h_row, start_column=p_h_col,
                   end_row=p_h_row + 1, end_column=p_h_col)
    ws.cell(column=p_h_col, row=p_h_row).value = "HORAS"
    ws.cell(column=p_h_col, row=p_h_row).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(column=p_h_col, row=p_h_row).fill = PatternFill('gray125')

    p_ce_per_col = p_h_col + 1
    p_ce_per_row = p_h_row
    ws.merge_cells(start_row=p_ce_per_row, start_column=p_ce_per_col,
                   end_row=p_ce_per_row + 1, end_column=p_ce_per_col)
    ws.cell(column=p_ce_per_col, row=p_ce_per_row).value = "% CE"
    ws.cell(column=p_ce_per_col, row=p_ce_per_row).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(column=p_ce_per_col, row=p_ce_per_row).fill = PatternFill('gray125')

    p_req_fe_col = p_ce_per_col + 1
    p_req_fe_row = p_h_row
    ws.merge_cells(start_row=p_req_fe_row, start_column=p_req_fe_col,
                   end_row=p_req_fe_row + 1, end_column=p_req_fe_col)
    ws.cell(column=p_req_fe_col, row=p_req_fe_row).value = "REQUISITO FE"
    ws.cell(column=p_req_fe_col, row=p_req_fe_row).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.cell(column=p_req_fe_col, row=p_req_fe_row).fill = PatternFill('gray125')
    ws.column_dimensions[p_req_fe_col_l].width = 15

    p_horas_dual_col = p_req_fe_col + 1
    p_horas_dual_row = p_req_fe_row
    ws.merge_cells(start_row=p_horas_dual_row, start_column=p_horas_dual_col,
                   end_row=p_horas_dual_row + 1, end_column=p_horas_dual_col)
    ws.cell(column=p_horas_dual_col, row=p_horas_dual_row).value = "HORAS DUAL"
    ws.cell(column=p_horas_dual_col, row=p_horas_dual_row).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.cell(column=p_horas_dual_col, row=p_horas_dual_row).fill = PatternFill('gray125')

    p_contenidos_col = p_horas_dual_col + 1
    p_contenidos_row = p_horas_dual_row
    ws.merge_cells(start_row=p_contenidos_row, start_column=p_contenidos_col,
                   end_row=p_contenidos_row + 1, end_column=p_contenidos_col)
    ws.cell(column=p_contenidos_col, row=p_contenidos_row).value = "CONTENIDOS"
    ws.cell(column=p_contenidos_col, row=p_contenidos_row).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.cell(column=p_contenidos_col, row=p_contenidos_row).fill = PatternFill('gray125')
    ws.column_dimensions[p_contenidos_col_l].width = 50

    p_complist_col = p_contenidos_col
    p_complist_row = p_contenidos_row - 6
    ws.cell(column=p_complist_col, row=p_complist_row).value = "OBJECTIUS"
    try:
        ws.cell(column=p_complist_col, row=p_complist_row + 1).value = str(modulo.ObjetivosGenerales)
    except Exception:
        print("  - INFO : No tiene Objetivos Generales")

    ws.cell(column=p_complist_col, row=p_complist_row + 2).value = "COMPETENCIES"
    try:
        ws.cell(column=p_complist_col, row=p_complist_row + 3).value = str(modulo.CompetenciasTitulo)
    except Exception:
        print("  - INFO : No tiene CompetenciasTitulo")

    p_ra_titulo_row = p_ra_titulo_row + 2
    ra_per = 100 / len(modulo.ResultadosAprendizaje) if modulo.ResultadosAprendizaje else 0

    for ra in modulo.ResultadosAprendizaje:
        listaCriterios = modulo.ResultadosAprendizaje[ra].CriteriosEvaluacion
        numCriterios = len(listaCriterios)

        ws.cell(column=p_ra_titulo_col, row=p_ra_titulo_row).value = ra + "." + modulo.ResultadosAprendizaje[ra].Resultado
        ws.cell(column=p_ra_titulo_col, row=p_ra_titulo_row).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells(start_row=p_ra_titulo_row, start_column=p_ra_titulo_col,
                       end_row=p_ra_titulo_row + numCriterios + 1, end_column=p_ra_titulo_col)

        ws.cell(column=p_ra_titulo_col + 1, row=p_ra_titulo_row).value = ra_per
        ws.cell(column=p_ra_titulo_col + 1, row=p_ra_titulo_row).style = numberStyle
        ws.cell(column=p_ra_titulo_col + 1, row=p_ra_titulo_row).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells(start_row=p_ra_titulo_row, start_column=p_ra_titulo_col + 1,
                       end_row=p_ra_titulo_row + numCriterios + 1, end_column=p_ra_titulo_col + 1)

        ws.merge_cells(start_row=p_ra_titulo_row, start_column=p_contenidos_col,
                       end_row=p_ra_titulo_row + numCriterios, end_column=p_contenidos_col)

        p_ce_row = p_ce_row + 2
        ws.cell(column=p_ce_col, row=p_ce_row).value = "TODOS"
        ws.cell(column=p_ce_col, row=p_ce_row).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(column=p_ce_col + 1, row=p_ce_row).value = "=SUM(F" + str(p_ce_row + 1) + ":F" + str(p_ce_row + numCriterios) + ")"
        ws.cell(column=p_ce_col + 2, row=p_ce_row).value = "=SUM(G" + str(p_ce_row + 1) + ":G" + str(p_ce_row + numCriterios) + ")"
        ws.cell(column=p_ce_col + 4, row=p_ce_row).value = "=SUM(I" + str(p_ce_row + 1) + ":I" + str(p_ce_row + numCriterios) + ")"

        ws.cell(column=p_ce_col, row=p_ce_row).fill = PatternFill('gray0625')
        ws.cell(column=p_ce_col + 1, row=p_ce_row).fill = PatternFill('gray0625')
        ws.cell(column=p_ce_col + 2, row=p_ce_row).fill = PatternFill('gray0625')
        ws.cell(column=p_ce_col + 4, row=p_ce_row).fill = PatternFill('gray0625')

        # Ingenieria para las competencias (same as json2excel.py)
        if numCriterios < 3:
            ws.cell(column=p_comp_col, row=p_ce_row + 1).value = "OBJECTIUS"
            ws.cell(column=p_comp_col, row=p_ce_row + 2).value = "COMPETENCIES"
        else:
            if numCriterios % 2 == 1:
                numCProf = numCriterios // 2 + 1
            else:
                numCProf = numCriterios // 2
            numCEmplea = numCriterios - numCProf

            ws.cell(column=p_comp_col, row=p_ce_row).value = "OBJECTIUS"
            ws.cell(column=p_comp_col, row=p_ce_row).fill = PatternFill('gray125')
            ws.merge_cells(start_row=p_ce_row + 1, start_column=p_comp_col,
                           end_row=p_ce_row + numCProf, end_column=p_comp_col)
            ws.cell(column=p_comp_col, row=p_ce_row + 1).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(column=p_comp_col, row=p_ce_row + numCProf + 1).value = "COMPETENCIES"
            ws.cell(column=p_comp_col, row=p_ce_row + numCProf + 1).fill = PatternFill('gray125')
            ws.merge_cells(start_row=p_ce_row + 2 + numCProf, start_column=p_comp_col,
                           end_row=p_ce_row + 2 + numCProf + numCEmplea - 1, end_column=p_comp_col)
            ws.cell(column=p_comp_col, row=p_ce_row + numCProf + 2).alignment = Alignment(horizontal='center', vertical='center')

        ce_per = 100 / numCriterios if numCriterios > 0 else 0
        for ce in listaCriterios:
            p_ce_row = p_ce_row + 1
            ws.cell(column=p_ce_col, row=p_ce_row).value = ce + ") " + listaCriterios[ce]
            ws.cell(column=p_ce_col, row=p_ce_row).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.cell(column=p_ce_col + 1, row=p_ce_row).value = 0
            ws.cell(column=p_ce_col + 2, row=p_ce_row).value = ce_per
            ws.cell(column=p_ce_col + 2, row=p_ce_row).style = numberStyle
            ws.cell(column=p_ce_col + 2, row=p_ce_row).alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

        p_ra_titulo_row = p_ra_titulo_row + numCriterios + 2

del wb['Sheet']
print(f" * Desant llibre optatives: {libro_path}")
wb.save(libro_path)

# -----------------------------------------------------------------------
# 2. Generate shared PDs
# -----------------------------------------------------------------------
def get_template_loader_and_env(template_name):
    searchpaths = [
        os.path.join(PROJECT_DIR, "templates"),
        os.path.join(PROJECT_DIR, "templates_INF"),
        os.path.join(PROJECT_DIR, "templates_SCO"),
    ]
    for sp in searchpaths:
        if os.path.exists(sp):
            loader = jinja2.FileSystemLoader(searchpath=sp)
            env = jinja2.Environment(loader=loader)
            try:
                env.get_template(template_name)
                return loader, env, sp
            except jinja2.TemplateNotFound:
                continue
    loader = jinja2.FileSystemLoader(searchpath=searchpaths[0])
    env = jinja2.Environment(loader=loader)
    return loader, env, searchpaths[0]


TEMPLATE_FILE = "PCCF_PD_Plantilla_MODULO_OPTATIVA.md"
tloader, tenv, used_path = get_template_loader_and_env(TEMPLATE_FILE)
print(f" * PD optatives: usant plantilles des de: {used_path}")
template = tenv.get_template(TEMPLATE_FILE)

for codigo, modulo in data_box.items():
    nom_net = modulo.nombre.replace(' ', '').replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
    fmod_borrador = os.path.join(plantilles_dir, f"PD_{codigo}_{nom_net}_BORRADOR.md")
    fmod_ok = os.path.join(plantilles_dir, f"PD_{codigo}_{nom_net}_OK.md")

    if os.path.exists(fmod_ok) or os.path.exists(fmod_borrador):
        existing = fmod_ok if os.path.exists(fmod_ok) else fmod_borrador
        print(f" PD optativa provisionat: {os.path.basename(existing)}")
        continue

    print(f" PD optativa generat: {os.path.basename(fmod_borrador)}")
    # Set synthetic OG/CPSS for template compatibility (placeholder data)
    modulo.OG = {k: k for k in (modulo.ObjetivosGenerales or [])}
    modulo.CPSS = {k: k for k in (modulo.CompetenciasTitulo or [])}
    outputText = template.render(modulo=modulo)
    with open(fmod_borrador, "w", encoding="utf-8") as f:
        f.write(outputText)

print(f" * PDs optatives generades a: {plantilles_dir}/")
