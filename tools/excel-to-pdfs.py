#!/usr/bin/env python3

import sys
import json
import os
import shutil
from openpyxl import load_workbook
import subprocess
import os
import warnings
warnings.filterwarnings('ignore')  # Por ahora ignoro todos los warnings

def exportar_rango_a_pdf(ruta_excel, hoja, rango, ruta_pdf):
    # Cargar el libro y la hoja
    wb = load_workbook(ruta_excel)
    try:
        ws = wb[hoja]
    except KeyError as ke:
        print(" * Hoja no encontrada : "+str(ke))
        return

    # Crear un nuevo libro y copiar solo el rango deseado
    nuevo_wb = wb.copy_worksheet(ws)
    for sheet in wb.worksheets:
        if sheet.title != hoja:
            wb.remove(sheet)

    # Eliminar filas y columnas fuera del rango
    cols = rango.split(":")[0][0], rango.split(":")[1][0]
    
    columna="I"
    ultima_fila = ws.max_row
    while ultima_fila > 0:
        if ws[f"{columna}{ultima_fila}"].value is not None:
            break
        ultima_fila -= 1

    # Me dejo esto apuntado por si acaso
    #filas = int(rango.split(":")[0][1:]), int(rango.split(":")[1][1:])
    filas = int(rango.split(":")[0][1:]), ultima_fila+10

    # Eliminar columnas no deseadas
    for col in range(ws.max_column, ord(cols[1]) - ord('A') + 1, -1):
        if chr(ord('A') + col - 1) > cols[1]:
            ws.delete_cols(col)
    for col in range(ord(cols[0]) - ord('A') + 1, 1, -1):
        if chr(ord('A') + col - 1) < cols[0]:
            ws.delete_cols(col)

    # Eliminar filas no deseadas
    for row in range(ws.max_row, filas[1], -1):
        ws.delete_rows(row)
    for row in range(filas[0] - 1, 1, -1):
        ws.delete_rows(row)

    # Guardar archivo temporal
    # Configurar la página en apaisado y ajustar al contenido
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    try:
        #ws.page_setup.fitToPage = True
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    except Exception as e :
        print(str(e))
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1  # Ajustar automáticamente la altura
    ws.page_margins.left = 0.1 
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.1
    ws.page_margins.bottom = 0.1
    ws.page_margins.header = 0.0
    ws.page_margins.footer = 0.0
    ruta_temp = "temp_rango.xlsx"
    wb.save(ruta_temp)

    # Convertir a PDF con LibreOffice
    cmd = [
        "libreoffice",
        "--headless",
        "--convert-to", "pdf",
        "--outdir", "/tmp",
        ruta_temp
    ]
    subprocess.run(cmd)

    # Mover el PDF generado a la ruta deseada
    pdf_generado = f"/tmp/{os.path.splitext(os.path.basename(ruta_temp))[0]}.pdf"
    os.rename(pdf_generado, ruta_pdf)

    # Limpiar archivo temporal
    os.remove(ruta_temp)


ruta_excel = sys.argv[1]
hoja = sys.argv[2]

from pccf_utils import get_hoja_label
hoja = get_hoja_label(hoja)
print(" Trabajando con la hoja : "+hoja)

rango = "B1:I64" 
ruta_pdf = "/tmp/cuadro-resumen.pdf"

exportar_rango_a_pdf(ruta_excel, hoja, rango, ruta_pdf)