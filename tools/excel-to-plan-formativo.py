#!/usr/bin/env python3

import sys
import json
import os
import shutil
from openpyxl import load_workbook
import subprocess
import os
import warnings
warnings.filterwarnings('ignore')  # Por ahora ignoro todos los warnings

fila_inicio=10
columna_a_buscar="I"

ruta_excel = sys.argv[1]
hoja = sys.argv[2]
plan_empresa_md = sys.argv[3]
hoja_orig=hoja

from pccf_utils import get_hoja_label
hoja = get_hoja_label(hoja)
print(" Trabajando con la hoja : "+hoja)

def generar_tabla_markdown(strings,hoja_orig):
    
    # Crear fichero y escribir
    with open(plan_empresa_md, 'a', encoding='utf-8') as fichero:
        
        print("## "+hoja_orig, file=fichero)
        print("\n\n", file=fichero)
        
        # Verificar si hay datos válidos (excluyendo "None.")
        datos_validos = []
        for s in strings:
            partes = s.split('.', 1)
            codigo = partes[0] + '.'
            if codigo != "None.":
                datos_validos.append(s)
        
        # Si no hay datos válidos, mostrar mensaje
        if not datos_validos:
            print("No s'ha definit cap RA d'este mòdul per a la Formació en empresa", file=fichero)
        else:
            # Generar tabla normal
            print("| RA       | Descripción |", file=fichero)
            print("|----------|-------------|", file=fichero)
            for s in datos_validos:
                # Separar el código (ej: "RA06.") del resto del texto
                partes = s.split('.', 1)
                codigo = partes[0]  # Sin punto final
                descripcion = partes[1] if len(partes) > 1 else ""
                print(f"| {codigo.ljust(8)} | {descripcion} |", file=fichero)
        
        print("\n\n\n",file=fichero)

# Ejemplo de uso:
strings = set()




wb = load_workbook(ruta_excel)
try:
    ws = wb[hoja]
except KeyError as ke:
    print(" * Hoja no encontrada : "+str(ke))
    sys.exit(0)

print(" * [ PCCF ] : Generando plan de Formacion en Empresa ")
for fila in range(fila_inicio, ws.max_row + 1):
    celda_actual = ws[f"{columna_a_buscar}{fila}"]
    # Comprobamos que el valor de la celda no es la suma simple
    # esto no es lo mejor, pero es que no se me ocurre otra cosa
    procesar_forzado=False
    if celda_actual.data_type == 'f' and "SUM" not in str(celda_actual.value):
        print("  [ Plan Formativo ] : Celda con formula expresada voluntariamente ")
        procesar_forzado=True
        
    if celda_actual.value is not None and celda_actual.data_type != 'f' or procesar_forzado:
        # Obtener la columna anterior (ej: 'B' si columna_a_buscar es 'C')
        columna_anterior = chr(ord(columna_a_buscar) - 7)
        celda_anterior = ws[f"{columna_anterior}{fila}"]

        # Verificar si la celda anterior está fusionada
        for merged_range in ws.merged_cells.ranges:
            if celda_anterior.coordinate in merged_range:
                # Obtener la celda superior izquierda del rango fusionado
                celda_fusionada = ws[merged_range.start_cell.coordinate]
                strings.add(str(celda_fusionada.value))
        strings.add(str(celda_anterior.value))
    else:
        pass

ras_ordenados=sorted(strings)
generar_tabla_markdown(ras_ordenados,hoja_orig)
