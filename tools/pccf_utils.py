# Modulo de Python para funciones comunes del PCCF

import os
import re
import json

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)

CICLES_INF = ["SMX", "DAM", "CEIABD", "FPBIIO"]
CICLES_SCO = ["APD", "EI", "IS"]
CICLES_CONEGUTS = sorted(CICLES_INF + CICLES_SCO, key=len, reverse=True)

OPTATIVES_PATH = os.path.join(PROJECT_DIR, "optatives", "optatives.json")
OPTATIVES_PLANTILLES = os.path.join(PROJECT_DIR, "optatives", "plantilles")

# Pattern per a noms de fitxer PD:
# PD_{CICLO}_{CODI}_{NOM}_{BORRADOR|OK}.md
PD_FILE_RE = re.compile(
    r'^PD_(' + '|'.join(CICLES_CONEGUTS) + r')_(\d+|[A-Z]+)_(.+?)_(BORRADOR|OK)\.md$'
)


def get_hoja_label(hoja):
    ## INFORMATICA
    # Comunes
    if hoja.startswith("Llenguatges de marques"): return "LM"
    if hoja.startswith("Sostenibilitat"): return "SOS"
    if hoja.startswith("Itinerari personal per a l'ocupabilitat II"): return "IPO2"
    if hoja.startswith("Itinerari personal per a l'ocupabilitat I"): return "IPO1"
    if hoja.startswith("Digitalització"): return "DIG"
    if hoja.startswith("Anglés Professional"): return "ANG"
    if hoja.startswith("Anglés oral"): return "AOEP"
    if hoja.startswith("Comunicació professional"): return "COM"
    if hoja.startswith("Projecte intermodular"): return "PI"
    if hoja.startswith("Introducció al Núvol"): return "NVL"

    if hoja.startswith("Muntatge"): return "MME"
    if hoja.startswith("Sistemes operatius mono"): return "SOM"
    if hoja.startswith("Aplicacions ofimàtiques"): return "AOF"
    if hoja.startswith("Sistemes operatius en xarxa"): return "SOX"
    if hoja.startswith("Seguretat informàtica"): return "SIN"
    if hoja.startswith("Serveis en xarxa"): return "SEX"
    if hoja.startswith("Aplicacions web"): return "AW"
    if hoja.startswith("Xarxes Locals"): return "XL"
    if hoja.startswith("Introducció a la Programació"): return "IPR"

    if hoja.startswith("Entorns de"): return "ED"
    if hoja.startswith("Sistemes Informàtics"): return "SI"
    if hoja.startswith("Bases de "): return "BBDD"
    if hoja.startswith("Programació de serveis i pro"): return "PSP"
    if hoja.startswith("Programació multimèdia i dispo"): return "PMDM"
    if hoja.startswith("Programació"): return "PRG"

    if hoja.startswith("Desenvolupament d'inter"): return "DI"
    if hoja.startswith("Accés a "): return "AD"
    if hoja.startswith("Sistemes de gestió empresarial"): return "SGE"

    if hoja.startswith("Models d"): return "MIA"
    if hoja.startswith("Sistemes d"): return "SAA"
    if hoja.startswith("Programació d"): return "PIA"
    if hoja.startswith("Sistemes de"): return "SBD"
    if hoja.startswith("Big Data"): return "BDA"

    if hoja.startswith("Montatge i manteniment"): return "MMEB"
    if hoja.startswith("Operacions auxiliars"): return "OA"
    if hoja.startswith("Ofimàtica"): return "OAD"
    if hoja.startswith("Instal·lació i manteniment"): return "IMXTD"
    if hoja.startswith("Ciències aplicades I"): return "CA1"
    if hoja.startswith("Ciències aplicades II"): return "CA2"
    if hoja.startswith("Comunicació i societat I"): return "CS1"
    if hoja.startswith("Comunicació i societat II"): return "CS2"

    ## SERVEIS A LA COMUNITAT
    if hoja.startswith("Organització de l'atenció"): return "OAPD"
    if hoja.startswith("Destreses socials"): return "DDSS"
    if hoja.startswith("Característiques i necessitats"): return "CNP"
    if hoja.startswith("Atenció i suport psicosocial"): return "ASP"
    if hoja.startswith("Suport a la comunicació"): return "SC"
    if hoja.startswith("Suport domiciliari"): return "SD"
    if hoja.startswith("Atenció sanitària"): return "AS"
    if hoja.startswith("Atenció higiènica"): return "AH"
    if hoja.startswith("Teleassistència"): return "TEL"
    if hoja.startswith("Primers auxilis"): return "PA"

    if hoja.startswith("Didàctica de l'educació infantil"): return "DEI"
    if hoja.startswith("Autonomia personal i salut infantil"): return "APSI"
    if hoja.startswith("El joc infantil"): return "JOC"
    if hoja.startswith("Expressió i comunicació"): return "EC"
    if hoja.startswith("Desenvolupament cognitiu"): return "DCM"
    if hoja.startswith("Desenvolupament socioafectiu"): return "DSA"
    if hoja.startswith("Habilitats socials"): return "HHSS"
    if hoja.startswith("Intervenció amb famílies"): return "IFAM"
    if hoja.startswith("Projecte d'atenció"): return "PAI"

    if hoja.startswith("Context de la intervenció social"): return "CIS"
    if hoja.startswith("Inserció sociolaboral"): return "ISL"
    if hoja.startswith("Atenció a les unitats de convivència"): return "AUC"
    if hoja.startswith("Mediació comunitària"): return "MC"
    if hoja.startswith("Suport a la intervenció educativa"): return "SIE"
    if hoja.startswith("Promoció de l'autonomia personal"): return "PAP"
    if hoja.startswith("Sistemes augmentatius"): return "SAAC"
    if hoja.startswith("Metodologia de la intervenció social"): return "MIS"

    return hoja


def get_familia(cicle):
    cicle = cicle.upper()
    if cicle in CICLES_INF:
        return "INF"
    if cicle in CICLES_SCO:
        return "SCO"
    return None


def get_moduls_del_cicle(cicle, familia=None):
    if familia is None:
        familia = get_familia(cicle)
    if familia is None:
        return {}
    json_path = os.path.join(PROJECT_DIR, f"boe_{familia}", f"rd-{cicle.lower()}.json")
    if not os.path.exists(json_path):
        return {}
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    return data.get("ModulosProfesionales", {})


def get_optatives(cicle=None, familia=None):
    """Carrega optatives/optatives.json i filtra per cicle/familia.

    Si cicle i familia es donen, retorna només els mòduls optatius que
    pertanyen a eixe cicle (segons el camp 'grups').
    Si no, retorna tots.
    """
    if not os.path.exists(OPTATIVES_PATH):
        return {}
    with open(OPTATIVES_PATH, encoding="utf-8") as f:
        optatives = json.load(f)

    if cicle is None or familia is None:
        return optatives

    cicle = cicle.upper()
    familia = familia.upper()
    result = {}
    for codi, modul in optatives.items():
        for g in modul.get("grups", []):
            if g.get("cicle", "").upper() == cicle and g.get("familia", "").upper() == familia:
                # Si hi ha codi alternatiu per a este cicle, usem-lo
                codi_real = modul.get("codis_alternatius", {}).get(cicle, codi)
                result[codi_real] = modul
                break
    return result


def get_moduls_merged(cicle, familia=None):
    """Fusiona els mòduls del cicle + optatives que pertanyen al cicle.

    Retorna un dict {codi: modul} on els codis alternatius es resolen
    automàticament.
    """
    moduls = get_moduls_del_cicle(cicle, familia)
    optatives = get_optatives(cicle, familia)
    # Les optatives sobreescriuen si hi ha conflicte (no n'hi hauria)
    moduls.update(optatives)
    return moduls


def parse_pd_filename(filename):
    """Parseja un nom de fitxer PD i retorna un dict o None."""
    if not filename.endswith(".md"):
        return None
    m = PD_FILE_RE.match(filename)
    if not m:
        return None
    return {
        "filename": filename,
        "ciclo": m.group(1),
        "codi": m.group(2),
        "nom": m.group(3),
        "estat": m.group(4),
    }


def check_excel_coherence(filepath):
    """Valida la coherència de l'Excel de pesos RA.

    Retorna una llista de missatges d'error (buida si tot correcte).
    """
    issues = []
    if not os.path.exists(filepath):
        return ["Excel no trobat: " + filepath]

    try:
        import openpyxl
    except ImportError:
        return []

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        return [f"Error en obrir Excel: {e}"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Buscar columna C (RA weights) i sumar
        total_weight = 0
        ra_count = 0
        errors = []
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=False):
            cell_b = row[1] if len(row) > 1 else None
            cell_c = row[2] if len(row) > 2 else None
            # Sols mirem files on la columna B comença per "RA" (descripció RA)
            if not (cell_b and cell_b.value and str(cell_b.value).strip().startswith("RA")):
                continue
            if cell_c and cell_c.value is not None:
                try:
                    val = float(cell_c.value)
                    total_weight += val
                    ra_count += 1
                except (ValueError, TypeError):
                    errors.append(f"  Fila {cell_c.row}: valor no numèric '{cell_c.value}'")

        if ra_count > 0:
            if abs(total_weight - 100) > 0.01:
                issues.append(
                    f"  Fulla '{sheet_name}': suma de pesos RA = {total_weight:.1f}% (hauria de ser 100%)"
                )
        else:
            issues.append(f"  Fulla '{sheet_name}': sense dades de pesos RA")

        for e in errors:
            issues.append(e)

    wb.close()
    return issues
